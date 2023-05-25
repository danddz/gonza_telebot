import telebot
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import os
import random

def creat_buttons(names_category, type_buttons='', name_menu='main_menu'):
    list_buttons = []
    tmp = []
    
    for i, value in enumerate(names_category):
        tmp.append(telebot.types.InlineKeyboardButton(value, callback_data=f'{name_menu}-{value}{type_buttons}'))
        if (i + 1) % 4 == 0 or i + 1 == len(names_category):
            list_buttons.append(tmp)
            tmp = []

    return list_buttons

def creat_answer_buttons(names_buttons, true_ans, name_category, type_buttons='-test', name_menu='main_menu'):
    list_buttons = []
    for value in names_buttons:
        list_buttons.append([telebot.types.InlineKeyboardButton(value, callback_data=f'{name_menu}-{name_category}-{value[:2]}{type_buttons}-{true_ans[:2]}')])
    return list_buttons

def creat_test_msg(list_dish, count_true_ans=0, count_false_ans=0):
    msg = 'Какое блюдо изображено на картинке?\n'
    
    for i, dish in enumerate(list_dish):
        msg += f'{i + 1}. {dish}\n'
    
    msg += f'\nПравильных ответов: {count_true_ans}\nНеправильных ответов: {count_false_ans}'
    return msg

def get_msg_test(data, main_menu, count_true_ans, count_false_ans):
    count = main_menu.get_count_rows(data) - 1
    numbers = [i + 1 for i in range(count)] 
    random.shuffle(numbers)

    numbers = numbers[:4]
    numbers_names = [main_menu.get_name_dish(data, i) for i in numbers]
    true_ans = numbers.index(random.choice(numbers))

    msg, image = main_menu.get_dish(data, numbers[true_ans])
    msg = creat_test_msg(numbers_names, count_true_ans, count_false_ans)

    list_buttons = ['Первое', 'Второе', 'Третье', 'Четвертое'][:len(numbers)]
    keyboard = telebot.types.InlineKeyboardMarkup(creat_answer_buttons(list_buttons, list_buttons[true_ans], data))
    
    return msg, image, keyboard

main_buttons = ['Меню', 'Бар', 'Тестирование']

main_menu_switch = [
    [
        telebot.types.InlineKeyboardButton('Предыдущее', callback_data='main_menu-Предыдущее'), 
        telebot.types.InlineKeyboardButton('Следующее', callback_data='main_menu-Следующее'), 
    ],
    [
        telebot.types.InlineKeyboardButton('Назад', callback_data='main_menu-Назад'), 
    ]
]

class main_menu_creater:
    def __init__(self, PATH):
        self.wb = openpyxl.load_workbook(PATH)
        self.list = self.wb.sheetnames
        self.image_PATH = 'files/images/main_menu'
        
    def get_count_rows(self, name_sheet):
        sheet = self.wb[name_sheet]
        max_row = 0
        while sheet.cell(row = max_row + 1, column = 1).value:
            max_row += 1
        return max_row
    
    def get_name_dish(self, name_category, number_dish):
        number_dish += 1
        return self.wb[name_category][f'A{number_dish}'].value
    
    def get_dish(self, name_category, number_dish):
        number_dish += 1
        
        name = self.wb[name_category][f'A{number_dish}'].value
        compound = self.wb[name_category][f'C{number_dish}'].value
        description = self.wb[name_category][f'D{number_dish}'].value
        gastropara = self.wb[name_category][f'E{number_dish}'].value
        
        count = self.get_count_rows(name_category) - 1
        
        ans =  f'''<i><b>{name}</b></i>
<b>Состав:</b>
{compound}
<b>Описание:</b>
{description}

<i>{gastropara}</i>
Блюдо №{number_dish - 1}/{count} ({name_category})
'''
        if len(ans) > 1000: # ЗАТЫЧКА !!!
            ans = ans[:300] + ans[800:]
        return ans, f'{self.image_PATH}/{name_category}/{number_dish - 1}.jpg'
    
    def creat_get_all_image_from_excel(self):
        if 'images' not in os.listdir(f'files'):
            os.makedirs(f'files/images')
                
        if 'main_menu' not in os.listdir(f'files/images'):
            os.makedirs(f'files/images/main_menu')
                
        for sheet in self.list:
            if sheet not in os.listdir(f'files/images/main_menu'):
                os.makedirs(f'files/images/main_menu/{sheet}')
                
            image_loader = SheetImageLoader(self.wb[sheet])
            
            for i in range(2, self.get_count_rows(sheet) + 1):    
                try:
                    image = image_loader.get(f'B{i}')
                except ValueError:
                    image = Image.open(f'files/no_image.jpg') 
                image.convert('RGB').save(f'files/images/main_menu/{sheet}/{i - 1}.jpg', optimize = True, quality = 10)

with open('files/token.txt', 'r', encoding='utf-8') as file:
    token = file.readline()
    
bot = telebot.TeleBot(token)
bot.set_my_commands([
        telebot.types.BotCommand('/start', 'Перезапуск бота'),
    ])

main_menu = main_menu_creater('files/menu.xlsx')

@bot.callback_query_handler(func=lambda call: True)
def brain_callback(query):
    message = query.message
    data = query.data
    text = message.caption

    if 'main_menu' in data:
        data = data.replace('main_menu-', '')
        if 'test' not in data:
            if data in main_menu.list:
                msg, image = main_menu.get_dish(data, 1)
                keyboard = telebot.types.InlineKeyboardMarkup(main_menu_switch)
                if msg != text:
                    bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(image, 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)

            elif data == 'Предыдущее':
                tmp = text[text.find('№') + 1:]
                index = int(tmp[:tmp.find('/')])
                count = int(tmp[tmp.find('/') + 1:tmp.find(' (')])
                category = tmp[tmp.find('(') + 1:tmp.find(')')]

                if index == 1:
                    index = count
                else:
                    index -= 1

                msg, image = main_menu.get_dish(category, index)
                keyboard = telebot.types.InlineKeyboardMarkup(main_menu_switch)
                if msg != text:
                    bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(image, 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)

            elif data == 'Следующее':
                tmp = text[text.find('№') + 1:]
                index = int(tmp[:tmp.find('/')])
                count = int(tmp[tmp.find('/') + 1:tmp.find(' (')])
                category = tmp[tmp.find('(') + 1:tmp.find(')')]

                if index == count:
                    index = 1
                else:
                    index += 1

                msg, image = main_menu.get_dish(category, index)
                keyboard = telebot.types.InlineKeyboardMarkup(main_menu_switch)
                if msg != text:
                    bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(image, 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)

            elif data == 'Назад':
                msg = 'Выберите тип блюда'
                keyboard = telebot.types.InlineKeyboardMarkup(creat_buttons(main_menu.list))
                bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(f'files/start_image.jpg', 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)
        else:
            data = data.replace('-test', '')
            if data in main_menu.list:
                
                msg, image, keyboard = get_msg_test(data, main_menu, 0, 0)
                
                if msg != text:
                    bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(image, 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)
            else:
                data = data.split('-')
                correct_ans = data[1] == data[2]
                
                count_true_ans = text[text.find('Правильных ответов: ') + len('Правильных ответов: '):]
                count_true_ans = int(count_true_ans[:count_true_ans.find('\n')]) + int(correct_ans)
                count_false_ans = int(text[text.find('Неправильных ответов: ') + len('Неправильных ответов: '):]) + int(not correct_ans)
                
                data = data[0]
                msg, image, keyboard = get_msg_test(data, main_menu, count_true_ans, count_false_ans)
                
                if msg != text:
                    bot.edit_message_media(media=telebot.types.InputMedia(type='photo', media=open(image, 'rb'), caption=msg, parse_mode='HTML'), chat_id=message.chat.id, message_id=message.message_id, reply_markup=keyboard)
                else:
                    print('esdf')
                
            
@bot.message_handler(content_types='text')
def message_reply(message):
    global main_menu
    
    text = message.text.lower()
    
    if text == '/start':
        msg = 'ПРИВЕТ'
        keyboard = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(*main_buttons)
        bot.send_message(message.chat.id, msg, reply_markup=keyboard)
    elif text == 'меню':
        msg = 'Выберите тип блюда'
        keyboard = telebot.types.InlineKeyboardMarkup(creat_buttons(main_menu.list))
        bot.send_photo(message.chat.id, photo=open(f'files/start_image.jpg', 'rb'), caption=msg, reply_markup=keyboard)
    
    elif text == 'тестирование':
        msg = 'Выберите тип блюда'
        keyboard = telebot.types.InlineKeyboardMarkup(creat_buttons(main_menu.list, type_buttons='-test'))
        bot.send_photo(message.chat.id, photo=open(f'files/start_image.jpg', 'rb'), caption=msg, reply_markup=keyboard)
    
    elif text == '/reload_info':
        main_menu = main_menu_creater('files/menu.xlsx')
        main_menu.creat_get_all_image_from_excel()

bot.polling()